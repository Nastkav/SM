import time

import openpyxl
from matplotlib import pyplot as plt

from create import Create
from process import Process
from model import Model


# res = list()
# task 1, 2
# c = Create(5)
#
# p1 = Process(5)
#
# p1.max_queue = 5
#
# c.distribution = 'exp'
# p1.distribution = 'exp'
#
# c.name = 'CREATE'
# p1.name = 'PROCESS-1'
#
# c.next_element = [p1]
#
# elements = [c, p1]
# model = Model(elements)
# res = model.simulate(1000)

# task3
# c = Create(5)
#
# p1 = Process(5, 2)
# p2 = Process(5)
# p3 = Process(5)
#
# c.next_element = [p1]
# p1.next_element = [p2]
# p2.next_element = [p3]
#
# p1.max_queue = 5
# p2.max_queue = 5
# p3.max_queue = 5
#
# c.distribution = 'exp'
# p1.distribution = 'exp'
# p2.distribution = 'exp'
# p3.distribution = 'exp'
#
# c.name = 'CREATE'
# p1.name = 'PROCESS-1'
# p2.name = 'PROCESS-2'
# p3.name = 'PROCESS-3'
#
# elements = [c, p1, p2, p3]
# model = Model(elements)
# res = model.simulate(1000)

# task6
# c = Create(5)
#
# p1 = Process(5)
# p2 = Process(5)
# p3 = Process(5)
#
# c.next_element = [p1]
# p1.next_element = [p2, p3]
# # p1.probability = ([0.7, 0.3])
# p1.priority = [2, 1]
# p1.max_queue = 5
# p2.max_queue = 5
# p3.max_queue = 5
#
# c.distribution = 'exp'
# p1.distribution = 'exp'
# p2.distribution = 'exp'
# p3.distribution = 'exp'
#
# c.name = 'CREATE'
# p1.name = 'PROCESS-1'
# p2.name = 'PROCESS-2'
# p3.name = 'PROCESS-3'
#
# elements = [c, p1, p2, p3]
# model = Model(elements)
# res = model.simulate(1000)

# task4:|

# workbook = openpyxl.Workbook()
# sheet = workbook.active
# tests = 10
#
# delay_createee = [5, 10, 5, 5, 5, 5, 1, 5, 5, 5]
# delay_process1 = [5, 5, 5, 5, 5, 5, 5, 1, 5, 5]
# delay_process2 = [5, 5, 5, 10, 5, 5, 5, 5, 1, 5]
# delay_process3 = [5, 5, 5, 5, 10, 5, 5, 5, 5, 1]
# channels_process1 = [1, 1, 3, 1, 1, 2, 3, 1, 1, 1, 1]
# channels_process2 = [1, 1, 3, 3, 1, 1, 2, 3, 1, 1, 1]
# channels_process3 = [1, 1, 3, 2, 3, 1, 1, 2, 3, 1, 1]
# max_queue1 = [5, 5, 5, 5, 5, 10, 5, 5, 5, 1, 5]
# max_queue2 = [5, 5, 5, 5, 5, 5, 10, 5, 5, 5, 1]
# max_queue3 = [5, 1, 5, 5, 5, 5, 5, 10, 5, 5, 5]
#
# column_names = [
#     'delay_create',
#     'channels_process1',
#     'max_queue1',
#     'delay_process1',
#     'channels_process2',
#     'max_queue2',
#     'delay_process2',
#     'channels_process3',
#     'max_queue3',
#     'delay_process3',
#     'quantity_create',
#     'quantity_p1',
#     'fail1',
#     'queue1',
#     'load1',
#     'quantity_p2',
#     'fail2',
#     'queue2',
#     'load2',
#     'quantity_p3',
#     'fail3',
#     'queue3',
#     'load3'
# ]
#
# for i in range(tests):
#     c = Create(delay_createee[i])
#
#     p1 = Process(delay_process1[i], channels_process1[i])
#     p2 = Process(delay_process2[i], channels_process2[i])
#     p3 = Process(delay_process3[i], channels_process3[i])
#
#     p1.max_queue = max_queue1[i]
#     p2.max_queue = max_queue2[i]
#     p3.max_queue = max_queue3[i]
#
#     c.distribution = 'exp'
#     p1.distribution = 'exp'
#     p2.distribution = 'exp'
#     p3.distribution = 'exp'
#
#     c.name = 'CREATOR'
#     p1.name = 'PROCESS-1'
#     p2.name = 'PROCESS-2'
#     p3.name = 'PROCESS-3'
#
#     c.next_element = [p1]
#     p1.next_element = [p2]
#     p2.next_element = [p3]
#
#     elements = [c, p1, p2, p3]
#     model = Model(elements)
#     res = model.simulate(1000)
#
#     result1 = {
#         'delay_createe': delay_createee[i],
#         'channels_process1': channels_process1[i],
#         'max_queue1': max_queue1[i],
#         'delay_process1': delay_process1[i],
#         'channels_process2': channels_process2[i],
#         'max_queue2': max_queue2[i],
#         'delay_process2': delay_process2[i],
#         'channels_process3': channels_process3[i],
#         'max_queue3': max_queue3[i],
#         'delay_process3': delay_process3[i],
#         'quantity_create':  res[0],
#         'quantity_p1': res[1],
#         'fail1': res[2][1],
#         'queue1': res[2][0],
#         'load1': res[2][2],
#         'quantity_p2': res[3],
#         'fail2': res[4][1],
#         'queue2': res[4][0],
#         'load2': res[4][2],
#         'quantity_p3': res[5],
#         'fail3': res[6][1],
#         'queue3': res[6][0],
#         'load3': res[6][2]
#     }
#
#
#     if i == 0:
#         for col_num, column_name in enumerate(column_names, start=1):
#             sheet.cell(row=1, column=col_num, value=column_name)
#
#     for col_num, value in enumerate(result1.values(), start=1):
#         sheet.cell(row=i + 2, column=col_num, value=value)
#
# workbook.save("RESULTS.xlsx")

# LAB4
def simple_model(N: int):
    c = Create(5)
    elements = [c]

    processes = []

    for i in range(N - 1):
        processes.append(Process(5))
        processes[i].max_queue = 100000
        processes[i].distribution = 'exp'
        processes[i].name = 'Process' + str(i + 1)
        if i > 0:
            processes[i - 1].next_element = [processes[i]]
        else:
            c.next_element = [processes[0]]

    elements.extend(processes)

    mod = Model(elements)
    return mod


def another_model(N: int):
    c = Create(5)
    elements = [c]

    processes1 = []
    for i in range(int(N / 2)):
        processes1.append(Process(5))
        processes1[i].max_queue = 100000
        processes1[i].distribution = 'exp'
        processes1[i].name = 'Process' + str(i + 1)

    processes2 = []
    for i in range(int(N / 2)):
        processes2.append(Process(5))
        processes2[i].max_queue = 100000
        processes2[i].distribution = 'exp'
        processes2[i].name = 'Process' + str(i + 1)

    c.next_element = [processes1[0], processes2[0]]
    for i in range(int(N / 2 - 1)):
        processes1[i].next_element = [processes1[i + 1]]
        processes2[i].next_element = [processes2[i + 1]]
    elements.extend(processes1)
    elements.extend(processes2)
    mod = Model(elements)
    return mod


def theor(V, timeMod=1, K=1):
    return V * timeMod * K


N = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]
tests = 5
results_time = []
theor_results_time = []

for j in range(len(N)):
    av_time = 0
    for i in range(tests):
        start = time.time()
        model = simple_model(N[j])
        # model = another_model(N[j])
        res = model.simulate(1000)
        end = time.time()
        av_time += end - start
    av_time = av_time / tests
    results_time.append(av_time)
    theor_results_time.append(theor(N[j]))
    print(results_time)

plt.title("Task 1")
plt.xlabel("Simple Model") #plt.xlabel("Another Model")
plt.ylabel("Time")
plt.plot(N, results_time, color="green")
plt.show()

plt.plot(N, theor_results_time, color="green")
plt.xlabel('N')
plt.ylabel('Time')
plt.title('Theoretical Model')
plt.show()




