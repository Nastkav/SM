import openpyxl
from create import Create
from process import Process
from model import Model

res = list()
# task 1, 2
# c = Create(5)
#
# p1 = Process(5)
#
# p1.max_queue = 5
#
# c.distribution = 'exp'
# p1.distribution = 'exp'
#
# c.name = 'CREATE'
# p1.name = 'PROCESS-1'
#
# c.next_element = [p1]
#
# elements = [c, p1]
# model = Model(elements)
# resu = model.simulate(1000)

# task3
# c = Create(5)
#
# p1 = Process(5)
# p2 = Process(5)
# p3 = Process(5)
#
# c.next_element = [p1]
# p1.next_element = [p2]
# p2.next_element = [p3]
#
# p1.max_queue = 5
# p2.max_queue = 5
# p3.max_queue = 5
#
# c.distribution = 'exp'
# p1.distribution = 'exp'
# p2.distribution = 'exp'
# p3.distribution = 'exp'
#
# c.name = 'CREATE'
# p1.name = 'PROCESS-1'
# p2.name = 'PROCESS-2'
# p3.name = 'PROCESS-3'
#
# elements = [c, p1, p2, p3]
# model = Model(elements)
# res = model.simulate(1000)

# task6
c = Create(5)

p1 = Process(5)
p2 = Process(5)
p3 = Process(5)

c.next_element = [p1]
p1.next_element = [p2, p3]
p1.probability = ([0.7, 0.3])

p1.max_queue = 5
p2.max_queue = 5
p3.max_queue = 5

c.distribution = 'exp'
p1.distribution = 'exp'
p2.distribution = 'exp'
p3.distribution = 'exp'

c.name = 'CREATE'
p1.name = 'PROCESS-1'
p2.name = 'PROCESS-2'
p3.name = 'PROCESS-3'

elements = [c, p1, p2, p3]
model = Model(elements)
res = model.simulate(1000)

# task4:|
#
# workbook = openpyxl.Workbook()
# sheet = workbook.active
# tests = 10
#
# delay_createee = [5, 10, 5, 5, 5, 5, 1, 5, 5, 5]
# delay_process1 = [5, 5, 5, 5, 5, 5, 5, 1, 5, 5]
# delay_process2 = [5, 5, 5, 10, 5, 5, 5, 5, 1, 5]
# delay_process3 = [5, 5, 5, 5, 10, 5, 5, 5, 5, 1]
# channels_process1 = [1, 1, 3, 1, 1, 2, 3, 1, 1, 1, 1]
# channels_process2 = [1, 1, 3, 3, 1, 1, 2, 3, 1, 1, 1]
# channels_process3 = [1, 1, 3, 2, 3, 1, 1, 2, 3, 1, 1]
# max_queue1 = [5, 5, 5, 5, 5, 10, 5, 5, 5, 1, 5]
# max_queue2 = [5, 5, 5, 5, 5, 5, 10, 5, 5, 5, 1]
# max_queue3 = [5, 1, 5, 5, 5, 5, 5, 10, 5, 5, 5]
#
# column_names = [
#     'delay_create',
#     'channels_process1',
#     'max_queue1',
#     'delay_process1',
#     'channels_process2',
#     'max_queue2',
#     'delay_process2',
#     'channels_process3',
#     'max_queue3',
#     'delay_process3',
#     'quantity_create',
#     'quantity_p1',
#     'fail1',
#     'queue1',
#     'load1',
#     'quantity_p2',
#     'fail2',
#     'queue2',
#     'load2',
#     'quantity_p3',
#     'fail3',
#     'queue3',
#     'load3'
# ]
#
# for i in range(tests):
#     c = Create(delay_createee[i])
#
#     p1 = Process(delay_process1[i], channels_process1[i])
#     p2 = Process(delay_process2[i], channels_process2[i])
#     p3 = Process(delay_process3[i], channels_process3[i])
#
#     p1.max_queue = max_queue1[i]
#     p2.max_queue = max_queue2[i]
#     p3.max_queue = max_queue3[i]
#
#     c.distribution = 'exp'
#     p1.distribution = 'exp'
#     p2.distribution = 'exp'
#     p3.distribution = 'exp'
#
#     c.name = 'CREATOR'
#     p1.name = 'PROCESS-1'
#     p2.name = 'PROCESS-2'
#     p3.name = 'PROCESS-3'
#
#     c.next_element = [p1]
#     p1.next_element = [p2]
#     p2.next_element = [p3]
#
#     elements = [c, p1, p2, p3]
#     model = Model(elements)
#     res = model.simulate(1000)
#
#     result1 = {
#         'delay_createe': delay_createee[i],
#         'channels_process1': channels_process1[i],
#         'max_queue1': max_queue1[i],
#         'delay_process1': delay_process1[i],
#         'channels_process2': channels_process2[i],
#         'max_queue2': max_queue2[i],
#         'delay_process2': delay_process2[i],
#         'channels_process3': channels_process3[i],
#         'max_queue3': max_queue3[i],
#         'delay_process3': delay_process3[i],
#         'quantity_create':  res[0],
#         'quantity_p1': res[1],
#         'fail1': res[2][1],
#         'queue1': res[2][0],
#         'load1': res[2][2],
#         'quantity_p2': res[3],
#         'fail2': res[4][1],
#         'queue2': res[4][0],
#         'load2': res[4][2],
#         'quantity_p3': res[5],
#         'fail3': res[6][1],
#         'queue3': res[6][0],
#         'load3': res[6][2]
#     }
#
#
#     if i == 0:
#         for col_num, column_name in enumerate(column_names, start=1):
#             sheet.cell(row=1, column=col_num, value=column_name)
#
#     for col_num, value in enumerate(result1.values(), start=1):
#         sheet.cell(row=i + 2, column=col_num, value=value)
#
# workbook.save("RESULTS.xlsx")
